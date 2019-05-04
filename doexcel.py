#!/usr/bin/python
# -*- coding: UTF-8 -*-

# desc
# keyKey1 = [0,6,7]
# keyKey2 = [0,3,7]
# 当各自上述的key对应值相等时，说明为同一条数据，否则，保存为两条数据
# 新的表的title为1，2表头所有字段。
# 如1表 表头为a,b,c，2表为d,e,f,g，生成新表表头为a,b,c,d,e,f,g


from openpyxl import Workbook
from openpyxl import load_workbook
import getopt
# wb = Workbook()

# ws = wb.active
# fileOne = load_workbook('D:\\ghosttusng\\RDEXcel\\1.xlsx')

import sys
reload(sys)
sys.setdefaultencoding("utf-8")

# opts, args = getopt.getopt(sys.argv[1:], "hi:o:") 
# for op, value in opts: 
#   if op == "-key1": 
#     # print('key1 = %s' % value)
#     print(value)
#   elif op == "-key2": 
#     # print('key2 = %s' % value)
#     print(value)
#   elif op == "-h": 
#     sys.exit()

fileName = r'D:/ghosttusng/RDExcel/1.xlsx'

title1 = []
title2 = []
newTitle = []
newSheetDatas = []

bankData1 = []
bankData2 = []

def initWb():
    wb = load_workbook(fileName)
    # print(wb.sheetnames)
    ws = wb.active
    return wb,ws

def readExcel():

    print(fileName)
    # inwb = openpyxl.load_workbook(fileName)
    wb = load_workbook(fileName)
    # print(wb.sheetnames)
    ws = wb.active

    wsNew = wb.get_sheet_by_name('new')
    print(wsNew)
    wb.remove(wsNew)
    wsNew = wb.create_sheet("new")
    

    for sheet in wb:
        print(sheet.title)

    rows = ws.max_row + 1
    cols = ws.max_column + 1

    for r in range(1, rows):
        arr = []
        for c in range(1, cols):
            arr.append(ws.cell(r, c).value)

        print(arr)
        wsNew.append(arr)
        

        print(r, ">>>>>>>>>>>>>>>>>>>")
        print("\r\n")

    wb.save(fileName)
    return

def getTitleAndDatas(ws):
    datas = []
    titles = []
    

    rows = ws.max_row + 1
    cols = ws.max_column + 1

    for c in range(1, cols):
        tv = ws.cell(1, c).value
        titles.append(tv)
        newTitle.append(tv)
    
    # for i in range(len(title1)):
    #     print("序号：%s   值：%s" % (i + 1, title1[i]))

    for r in range(2, rows):
        rowData = []
        dic = {}
        for c in range(1, cols):
            # dic = {}
            title = titles[c-1]
            print("index:%s  title:%s value:%s" % (c,title,ws.cell(r, c).value))
            dic[title] = ws.cell(r, c).value
            # print(dic)
            # rowData.append(dic)
        # if r > 3:
        #     break
        print(r, ">>>>>>>>>>>>>>>>>>>")
        datas.append(dic)

    return titles,datas

# 先获取两个sheet页面中的值，存入arr中，其中每个值又是一个dict
def do():
    wb,ws = initWb()

    ws1 = wb.get_sheet_by_name('1')
    ws2 = wb.get_sheet_by_name('2')

    title1,datas1 = getTitleAndDatas(ws1)
    for g in range(len(title1)):
        bankData1.append('')

    title2,datas2 = getTitleAndDatas(ws2)
    for d in range(len(title2)):
        bankData2.append('')
    # print('表2数据：========')
    # for o in range(len(datas2)):
    #     print(datas2[o])
    newSheetDatas.append(newTitle)

    # print('表一表头：========')
    # for o in range(len(title1)):
    #     print(title1[o])

    # print('表2表头：========')
    # for j in range(len(title2)):
    #     print(title2[j])
    # print('========')

    # 以数据少的为外层遍历，遍历完毕时如果有剩余，则继续添加新的数据到新表
    datas1Len = len(datas1)
    datas2Len = len(datas2)
    print('表1数据条数:%s,表2数据条数:%s' % (datas1Len,datas2Len))
    pacakgeDatas(datas1,datas2,title1,title2,wb)
    
    return

def pacakgeDatas(datas1,datas2,sheetTitle1,sheetTitle2,wb):
    datas1Len = len(datas1)
    datas2Len = len(datas2)
    lenth = datas1Len
    leftLenth = datas2Len - datas1Len
    leftDatasFlag = 0 # 0-表示短的数据是datas1,1-表示短的数据是datas2
    # if datas1Len > datas2Len :
    #     lenth = datas2Len
    #     leftLenth = datas1Len - datas2Len
    #     leftDatasFlag = 1
    
    for i in range(lenth):
        rowDic1 = datas1[i]
        
        # keyKey1 = [0,6,7]
        # keyKey2 = [0,3,7]
        key1Code =  sheetTitle1[0]
        key1CustomName =  sheetTitle1[6]
        key1Num =  sheetTitle1[7]

        for sheet2idx in range(datas2Len):
            rowDic2 = datas2[sheet2idx]
            key2Code =  sheetTitle2[0]
            key2CustomName =  sheetTitle2[3]
            key2Num =  sheetTitle2[7]

            if rowDic1[key1Code] == rowDic2[key2Code] and rowDic1[key1CustomName] == rowDic2[key2CustomName] and rowDic1[key1Num] == rowDic2[key2Num]:
                print('index of the same data in EXCEL :%d' % (i+2))
                datas = []
                for j in range(len(sheetTitle1)):
                    key = sheetTitle1[j]
                    datas.append(rowDic1[key])
                # for key in rowDic1:
                #     datas.append(rowDic1[key])
                # for key in rowDic2:
                for j0 in range(len(sheetTitle2)):
                    key = sheetTitle2[j0]
                    datas.append(rowDic2[key])
                newSheetDatas.append(datas)
            else:
                print('index of the different data in EXCEL:%d' % (i+2))
                datas = []
                for j in range(len(sheetTitle1)):
                    key = sheetTitle1[j]
                    datas.append(rowDic1[key])

                for p in range(len(bankData2)):
                    datas.append(bankData2[p])
                newSheetDatas.append(datas)
                

                datasT = []
                for p in range(len(bankData1)):
                    datasT.append(bankData1[p])
                for j0 in range(len(sheetTitle2)):
                    key = sheetTitle2[j0]
                    datasT.append(rowDic2[key])
                newSheetDatas.append(datasT)
                
            print('========')
    
    # if leftDatasFlag == 0:
    #     for k in range(leftLenth):
    #         datas = []
    #         print('正常序号:%d,剩余序号:%d' % (k,lenth+k))
    #         for p in range(len(bankData2)):
    #             datas.append(bankData2[p])
            
    #         leftDatas = datas2[lenth+k]
    #         # for key in leftDatas:
    #         for j0 in range(len(sheetTitle2)):
    #             key = sheetTitle2[j0]
    #             datas.append(leftDatas[key])

    #         newSheetDatas.append(datas)
    # else:
    #     for idx in range(leftLenth):
    #         datas = []
    #         print('正常序号:%d,剩余序号:%d' % (k,lenth+idx))
    #         leftDatas = datas1[lenth+idx]
    #         # for key in leftDatas:
    #         for j0 in range(len(sheetTitle2)):
    #             key = sheetTitle2[j0]
    #             datas.append(leftDatas[key])

    #         for p in range(len(bankData1)):
    #             datas.append(bankData1[p])

    #         newSheetDatas.append(datas)
    
    # print(newSheetDatas)
    appendToSheet(newSheetDatas,wb)
    return

def pacakgeDatasForSheet(keyCode,sheetTitle1,sheetTitle2):
    arr = []
    for i in range(len(newTitle)):
        if keyCode == newTitle[i]:
            if  i > 0:
                # 前面的值全为空
                for j in range(len(sheetTitle1)):
                    arr.append('')
                break
            else:
                # 后面的值全为空
                for k in range(len(sheetTitle2)):
                    arr.append('')
                break
    return arr

# 把数据添加到新表中
def appendToSheet(datas,wb):
    wsNew = wb.get_sheet_by_name('new')
    wb.remove(wsNew)
    wsNew = wb.create_sheet("new")

    for k in range(len(datas)):
        # print('存入数据===================')
        # print(datas[k])
        wsNew.append(datas[k])
    wb.save(fileName)
    return

def sortedDictValues(adict):
    keys = adict.keys()
    keys.sort()
    return map(adict.get, keys)

# readExcel()
do()

