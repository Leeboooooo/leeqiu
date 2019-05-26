#!/usr/bin/python
# -*- coding: UTF-8 -*-

# desc
# keyKey1 = [0,6,7]
# keyKey2 = [0,3,7]
# 当各自上述的key对应值相等时，说明为同一条数据，否则，保存为两条数据
# 新的表的title为1，2表头所有字段。
# 如1表 表头为a,b,c，2表为d,e,f,g，生成新表表头为a,b,c,d,e,f,g


from openpyxl import Workbook
from openpyxl import load_workbook
import getopt
# wb = Workbook()

# ws = wb.active
# fileOne = load_workbook('D:\\ghosttusng\\RDEXcel\\1.xlsx')

import sys
reload(sys)
sys.setdefaultencoding("utf-8")

# opts, args = getopt.getopt(sys.argv[1:], "hi:o:") 
# for op, value in opts: 
#   if op == "-key1": 
#     # print('key1 = %s' % value)
#     print(value)
#   elif op == "-key2": 
#     # print('key2 = %s' % value)
#     print(value)
#   elif op == "-h": 
#     sys.exit()

fileName = r'1.xlsx'

title1 = []
title2 = []
newTitle = []
newSheetDatas = []

bankData1 = []
bankData2 = []

def initWb(fileName):
    wb = load_workbook(fileName)
    # print(wb.sheetnames)
    ws = wb.active
    return wb,ws

def readExcel():

    print(fileName)
    # inwb = openpyxl.load_workbook(fileName)
    wb = load_workbook(fileName)
    # print(wb.sheetnames)
    ws = wb.active

    wsNew = wb.get_sheet_by_name('new')
    print(wsNew)
    wb.remove(wsNew)
    wsNew = wb.create_sheet("new")
    

    # for sheet in wb:
        # print(sheet.title)

    rows = ws.max_row + 1
    cols = ws.max_column + 1

    for r in range(1, rows):
        arr = []
        for c in range(1, cols):
            arr.append(ws.cell(r, c).value)

        # print(arr)
        wsNew.append(arr)
        

        # print(r, ">>>>>>>>>>>>>>>>>>>")
        # print("\r\n")

    wb.save(fileName)
    return

def getTitleAndDatas(ws):
    datas = []
    titles = []
    

    rows = ws.max_row + 1
    cols = ws.max_column + 1

    for c in range(1, cols):
        tv = ws.cell(1, c).value
        titles.append(tv)
        newTitle.append(tv)
    
    # for i in range(len(title1)):
    #     print("序号：%s   值：%s" % (i + 1, title1[i]))

    for r in range(2, rows):
        rowData = []
        dic = {}
        for c in range(1, cols):
            # dic = {}
            title = titles[c-1]
            # print("index:%s  title:%s value:%s" % (c,title,ws.cell(r, c).value))
            dic[title] = ws.cell(r, c).value
            # print(dic)
            # rowData.append(dic)
        # if r > 3:
        #     break
        # print(r, ">>>>>>>>>>>>>>>>>>>")
        datas.append(dic)

    return titles,datas

# 先获取两个sheet页面中的值，存入arr中，其中每个值又是一个dict
def do(fileName):
    wb,ws = initWb(fileName)

    ws1 = wb.get_sheet_by_name('发货明细')
    ws2 = wb.get_sheet_by_name('销售明细')

    title1,datas1 = getTitleAndDatas(ws1)
    for g in range(len(title1)):
        bankData1.append('')

    title2,datas2 = getTitleAndDatas(ws2)
    for d in range(len(title2)):
        bankData2.append('')
    # print('表2数据：========')
    # for o in range(len(datas2)):
    #     print(datas2[o])
    newSheetDatas.append(newTitle)

    # print('表一表头：========')
    # for o in range(len(title1)):
    #     print(title1[o])

    # print('表2表头：========')
    # for j in range(len(title2)):
    #     print(title2[j])
    # print('========')

    # 以数据少的为外层遍历，遍历完毕时如果有剩余，则继续添加新的数据到新表
    # datas1Len = len(datas1)
    # datas2Len = len(datas2)
    # print('表1数据条数:%s,表2数据条数:%s' % (datas1Len,datas2Len))
    pacakgeDatas(datas1,datas2,title1,title2,wb)
    
    return

def pacakgeDatas(datas1,datas2,sheetTitle1,sheetTitle2,wb):
    datas1Len = len(datas1)
    datas2Len = len(datas2)
    # print('第一页有 %d 条，第二页有 %d 条' % (datas1Len,datas2Len))
    sameDatas1 = []
    sameDatas2 = []
    for i in range(datas1Len):
        rowDic1 = datas1[i]
        key1Code =  sheetTitle1[0]
        key1CustomName =  sheetTitle1[6]
        key1Num =  sheetTitle1[7]
        isRecord = 0 #1-表示左表的值已经出现过对应的
        for sheet2idx in range(datas2Len):
            rowDic2 = datas2[sheet2idx]
            key2Code =  sheetTitle2[0]
            key2CustomName =  sheetTitle2[3]
            key2Num =  sheetTitle2[7]
            
            if rowDic1[key1Code] == rowDic2[key2Code] and rowDic1[key1CustomName] == rowDic2[key2CustomName] and rowDic1[key1Num] == rowDic2[key2Num]:
                # print('\n左边表 [%d] 相同值标记前:%d' % (i+2,isRecord))
                if isRecord == 0:
                    isR = 'isR' in rowDic2.keys()
                    # print(isR)
                    if isR:
                        # print('右边表[%d] isR = %d' % (sheet2idx+2,rowDic2['isR']))
                        # print('\n!!!!!!!!!!!!!!!!!!!!!!!!!!!')
                        # print(datas2[sheet2idx])
                        # print('右边表[%d]' % (sheet2idx+2))
                        # print('!!!!!!!!!!!!!!!!!!!!!!!!!!!\n')
                        continue
                    else:
                        # print('index of the same data in EXCEL1 :%d,index of 2:%d' % (i+2,sheet2idx+2))
                        print(' --------------- Progressing %.2f%%  --------------- ' % (i/float(datas1Len)*100))
                        # print('key1Code:%s,value1:%s,key1CustomName:%s,valueCustom:%s,key1Num:%s,valueNum:%s' % (key1Code,rowDic1[key1Code],key1CustomName,rowDic1[key1CustomName],key1Num,rowDic1[key1Num]))
                        # print('key2Code:%s,value2:%s,key2CustomName:%s,valueCustom2:%s,key2Num:%s,valueNum2:%s' % (key2Code,rowDic2[key2Code],key2CustomName,rowDic2[key2CustomName],key2Num,rowDic2[key2Num]))
                        datas = []
                        for j in range(len(sheetTitle1)):
                            key = sheetTitle1[j]
                            datas.append(rowDic1[key])

                        for j0 in range(len(sheetTitle2)):
                            key = sheetTitle2[j0]
                            datas.append(rowDic2[key])

                        sameDatas1.append(rowDic1)
                        sameDatas2.append(rowDic2)
                        newSheetDatas.append(datas)
                        isRecord = 1
                        rowDic2['isR'] = 1
                        # print('\n-----------------------------------')
                        # print('右边表[%d]' % (sheet2idx+2))
                        # print(datas2[sheet2idx])
                        # print('-----------------------------------\n')
                        continue
                        
                # print('左边表 [%d] 相同值标记:%d' % (i+2,isRecord))
                
            
            # print(',.,..,.,\n')
            # print(datas2[sheet2idx])


    # print('====第1页相同数据====\r\n')
    # print(len(sameDatas1))
    # print('====第1页相同数据====\r\n')

    # print('====第2页相同数据====\r\n')
    # print(len(sameDatas2))
    # print('====第2页相同数据====\r\n')

    

    pacakgeDiffDatas(datas1,sameDatas1,sheetTitle1,1)
    # newSheetDatas.append(datasLeft)
    pacakgeDiffDatas(datas2,sameDatas2,sheetTitle2,0)
    # newSheetDatas.append(datasRight)

    # print('>>>>>存入sheet的数据<<<<<<\r\n')
    # print(newSheetDatas)
    appendToSheet(newSheetDatas,wb)
    return

# 处理两页不同的数据。拼接到sameDatas后形成新的sheetDatas
def pacakgeDiffDatas(datas,sameDatas,sheetTitle,left):
    # print('总的表：\r\n')
    # print(datas)
    # print('相同数据的表：\r\n')
    # print(sameDatas)
    for i in range(len(sameDatas)):
        # print('\n对比的数据[%d]' % i)
        # print(sameDatas[i])
        index = datas.index(sameDatas[i])
        datas.pop(index)

    for j in range(len(datas)):
        d = []
        if left == 1 :
            for key in sheetTitle:
                d.append(datas[j][key])
            for p in range(len(bankData2)):
                d.append(bankData2[p])
        else:
            for p in range(len(bankData1)):
                d.append(bankData1[p])
            for key in sheetTitle:
                d.append(datas[j][key])
        newSheetDatas.append(d)
    # return sheetDatas

def pacakgeDatasForSheet(keyCode,sheetTitle1,sheetTitle2):
    arr = []
    for i in range(len(newTitle)):
        if keyCode == newTitle[i]:
            if  i > 0:
                # 前面的值全为空
                for j in range(len(sheetTitle1)):
                    arr.append('')
                break
            else:
                # 后面的值全为空
                for k in range(len(sheetTitle2)):
                    arr.append('')
                break
    return arr

# 把数据添加到新表中
def appendToSheet(datas,wb):
    wsNew = wb.get_sheet_by_name('new')
    wb.remove(wsNew)
    wsNew = wb.create_sheet("new")

    for k in range(len(datas)):
        # print('存入数据===================')
        # print(datas[k])
        wsNew.append(datas[k])
    wb.save(fileName)
    print(' --------------- Progressing 100% --------------- ')
    print(' --------------- DONE! --------------- ')
    return

def sortedDictValues(adict):
    keys = adict.keys()
    keys.sort()
    return map(adict.get, keys)

def test():
    if len(sys.argv) < 2:
        print('兄弟，输入要对比的文件地址，比如E://RDEXcel//1.xlsx')
        return
    fileName = sys.argv[1] + '.xlsx'
    print(fileName)
    do(fileName)
# readExcel()
do('1.xlsx')
# test()
